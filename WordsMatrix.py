#!/usr/bin/env python
# -*- coding: utf-8 -*-

import os, io
import sys, getopt
import random

from openpyxl import load_workbook


# input: words list
# output 
#   1. random matrix view with colume and row parameter
# 
# 

def MatrixFromXLS(filepath, sel_sheets, river):
    print filepath
    wb = load_workbook(filename=filepath, read_only=True, data_only=True)
    worksheets = wb.get_sheet_names()
    print worksheets
    words = {}
    formatstr = ''
    print sorted(worksheets)
    for sheetname in sorted(worksheets):
        print sheetname
        iSskip = True
        for sheetsel in sel_sheets:
            if sheetsel is not None and sheetname == sheetsel:
                iSskip = False
                break
        
        if iSskip:
            continue

        ws = wb[sheetname]
        words[sheetname] = []
        formatstr = '%%-%ds' % river
        for row in ws.rows:
            if len(row) > 1:
                if row[1].value == None:
                    if row[0].value != None:
                        words[sheetname].append(formatstr % row[0].value)
                else:
                    if isinstance(row[1].value,float):
                        words[sheetname].append(formatstr % str(row[0].value))
                    else:
                        words[sheetname].append(formatstr % row[1].value)
            else:
                if row[0].value != None:
                        words[sheetname].append(formatstr % row[0].value)
            #'%-16s'
    return words


def MatrixFromRawTxt(inputfile):
    content = []
    with open(inputfile) as f:
        for line in f.readlines():
            word = line.strip().rstrip('\n')
            if  word != '':
                content.append(word)
       
    random.shuffle(content)
    print {'allinone':content}

def CreateMatrixFile(content, outputfile, columns):
    with io.open(outputfile, 'w', encoding='utf8') as of:
        for key in content:
            k = 1
            wordline = ''
            of.writelines(u'\n')
            of.writelines(key.decode('ascii'))
            of.writelines(u'\n')
            for word in content[key]:
                
                if k % columns == 0:
                    wordline += word.strip()
                    wordline +=u'\n'
                    of.writelines(wordline)
                    wordline = ''
                else:
                    wordline += word

                k+=1
            if wordline != '':
                print wordline


if __name__ == '__main__':

    inputfile = ''
    outputfile = ''
    columns = 5
    river = 16
    sheetnames = []
    try:
        opts, args = getopt.getopt(sys.argv[1:],"hi:o:c:s:r:",["ifile=","ofile=","columns=","sheet=","river="])
    except getopt.GetoptError:
        print 'test.py -i <inputfile> -o <outputfile>'
        sys.exit(2)
    for opt, arg in opts:
        if opt == '-h':
            print 'test.py -i <inputfile> -s <xls sheet name> -o <outputfile> -c <numbers>'
            sys.exit()
        elif opt in ("-i", "--ifile"):
            inputfile = arg
        elif opt in ("-o", "--ofile"):
            outputfile = arg
        elif opt in ("-c", "--columns"):
            columns = int(arg)
        elif opt in ("-s","--sheet"):
            sheetnames = arg.split(",")
        elif opt in ("-r","--river"):
            river = int(arg)

    print 'Input file is "', inputfile
    print 'Output file is "', outputfile
    print sheetnames
    dat = MatrixFromXLS(inputfile, sheetnames,river)
    #MatrixFromRawTxt(inputfile,outputfile,columns)
    CreateMatrixFile(dat,outputfile,columns)

   